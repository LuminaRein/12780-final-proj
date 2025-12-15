import json
import csv
from pathlib import Path
from io import BytesIO
from django.http import JsonResponse, HttpResponse
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils.timezone import localtime
from django.conf import settings
from django.shortcuts import render

import openpyxl
from openpyxl.utils import get_column_letter

from .models import HexPoint, Shape, ShapePoint

def home(request):
    return render(request, "home.html")

def _get_or_create_hexpoint(q, r, s):
    q = int(q); r = int(r); s = int(s)
    if (q + r + s) != 0:
        raise ValidationError("Cube coords sum should be 0.")
    obj, _ = HexPoint.objects.get_or_create(q=q, r=r, s=s, defaults={"name": ""})
    return obj


def _read_json(request):
    try:
        return json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        raise ValidationError("Invalid JSON body")


def create_circle(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)

    data = _read_json(request)
    origin = data.get("origin", None)
    if origin is None:
        return JsonResponse({"ok": False, "error": "Missing origin"}, status=400)

    try:
        magnitude = int(data.get("magnitude"))
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid magnitude"}, status=400)

    points = data.get("points", [])

    try:
        with transaction.atomic():
            o = _get_or_create_hexpoint(origin["q"], origin["r"], origin["s"])
            shape = Shape.objects.create(shape_type="circle", origin=o, magnitude=magnitude)

            for i, p in enumerate(points):
                hp = _get_or_create_hexpoint(p["q"], p["r"], p["s"])
                ShapePoint.objects.create(shape=shape, point=hp, idx=i)

        return JsonResponse({"ok": True, "shape_id": shape.id})
    except ValidationError as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)


def create_triangle(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)

    data = _read_json(request)
    vertices = data.get("vertices", None)
    if not isinstance(vertices, list) or len(vertices) != 3:
        return JsonResponse({"ok": False, "error": "Triangle requires 3 clicks"}, status=400)

    try:
        magnitude = int(data.get("magnitude", 0))
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid magnitude"}, status=400)

    points = data.get("points", None)

    try:
        with transaction.atomic():
            v0 = vertices[0]
            origin = _get_or_create_hexpoint(v0["q"], v0["r"], v0["s"])
            shape = Shape.objects.create(shape_type="triangle", origin=origin, magnitude=magnitude)

            for i, v in enumerate(vertices):
                hp = _get_or_create_hexpoint(v["q"], v["r"], v["s"])
                ShapePoint.objects.create(shape=shape, point=hp, idx=i)

            if points is not None:
                start = 3
                for j, p in enumerate(points):
                    hp = _get_or_create_hexpoint(p["q"], p["r"], p["s"])
                    ShapePoint.objects.create(shape=shape, point=hp, idx=start + j)

        return JsonResponse({"ok": True, "shape_id": shape.id})
    except ValidationError as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)


def export_shapes_xlsx(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shapes"

    headers = ["id", "shape_type", "origin_id", "origin_q", "origin_r", "origin_s", "magnitude", "created_at"]
    ws.append(headers)

    qs = Shape.objects.select_related("origin").order_by("id")
    for sh in qs:
        o = sh.origin
        ws.append([
            sh.id,
            sh.shape_type,
            o.id if o else None,
            o.q if o else None,
            o.r if o else None,
            o.s if o else None,
            sh.magnitude,
            localtime(sh.created_at).replace(tzinfo=None) if sh.created_at else None,
        ])

    ws.freeze_panes = "A2"

    for col_idx, h in enumerate(headers, start=1):
        max_len = max(len(str(h)), *(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(40, max(10, max_len + 2))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="shapes.xlsx"'
    return resp

def clear_db(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)

    with transaction.atomic():
        ShapePoint.objects.all().delete()
        Shape.objects.all().delete()
        HexPoint.objects.all().delete()

    return JsonResponse({"ok": True})
